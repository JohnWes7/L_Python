# pip install -i https://pypi.tuna.tsinghua.edu.cn/simple

def open():
    from openpyxl import load_workbook
    wb = load_workbook("data/test.xlsx")
    sh1 = wb.active  # 激活默认第一个
    print(wb.sheetnames)
    sh2 = wb["Sheet1"]
    print("done")


def show_sheets():
    from openpyxl import load_workbook
    wb = load_workbook("data/test.xlsx")
    print(wb.sheetnames)
    for sh in wb:
        print(sh.title)


def get_one_value():
    from openpyxl import load_workbook
    wb = load_workbook("data/test.xlsx")
    sheet = wb[wb.sheetnames[0]]
    cell = sheet.cell(5, 2)
    cell2 = sheet["B3"]
    cell.value
    cell2.value
    print(f"{cell} : {cell.value}")
    print(f"{cell2} : {cell2.value}")

    for row in sheet.rows: #还有个sheet.columns
        for cell in row:
            print(cell.value, end="\t")
        print()
    print()

    part = sheet["A1":"C3"]
    for row in part:
        for cell in row:
            print(cell.value,end="\t")
        print()
    print()

# open()
# show_sheets()


get_one_value()
