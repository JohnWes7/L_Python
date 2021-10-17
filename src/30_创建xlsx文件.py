from openpyxl.reader.excel import load_workbook


def creat_xlsx():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh2 = wb.create_sheet("Info")
    wb.save("data/test_creat_xlsx.xlsx")


def set_value():
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb[wb.sheetnames[0]]
    sheet.title = "Info"

    sheet["A1"] = "hello"
    sheet["A2"] = "python"
    sheet["B1"] = "excel"

    wb.save("data/test_creat_xlsx.xlsx")  # 如果文件存在会直接覆盖，如果不存在就会创建
    print("done")


def set_values():
    rows = [["index", "name", "price"],
            [1, "Cola", 10],
            [2, "beer", 12],
            [3, "water", 2]]

    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb[wb.sheetnames[0]]

    for row in rows:
        sheet.append()

    wb.save("data/test_creat_xlsx.xlsx")


def practice1():
    data = [["name", "rank", "type", "country"],
            ["Laffey", 4, "DD", "USS"],
            ["Essex", 5, "CV", "USS"],
            ["Unicorn", 4, "CVL", "HMS"],
            ["Amagi", 5, "CC", "IJN"],
            ["Odin", 5, "CC", "KMS"]]

    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb[wb.sheetnames[0]]

    for row in data:
        sheet.append(row)

    wb.save("data/practice1_azur_lane.xlsx")
    print("done")


def practice2():
    import openpyxl
    wb = load_workbook("data/practice1_azur_lane.xlsx")
    sheet = wb[wb.sheetnames[0]]

    # name_list
    # rank_list
    # type_list
    # country_list

    for c in sheet.columns:
        if c[0].value == "name":
            name_list = c
        if c[0].value == "rank":
            rank_list = c
        if c[0].value == "type":
            type_list = c
        if c[0].value == "country":
            country_list = c

    count_c = 0
    out_list = []
    for i in range(len(country_list)):
        if country_list[i].value == "USS":
            count_c += 1
            out_list.append(name_list[i].value)

    print(f"country == USS | count : {count_c} : {out_list}")

    count_c = 0
    out_list.clear()
    for i in range(len(type_list)):
        if type_list[i].value == "CC":
            count_c += 1
            out_list.append(name_list[i].value)

    print(f"type == CC | count : {count_c} : {out_list}")


if __name__ == "__main__":
    # creat_xlsx()
    # set_value()
    # set_values()
    # practice1()
    practice2()
