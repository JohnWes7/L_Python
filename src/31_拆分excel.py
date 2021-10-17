from config import config
import os
import openpyxl

file_name = "/practice1_azur_lane.xlsx"
data_path = config.absolute_data_PATH + "/practice1_azur_lane.xlsx"
creat_path = config.absolute_create_files_PATH + "/practice1_azur_lane.xlsx"

wb = openpyxl.load_workbook(data_path)
sheet = wb[wb.sheetnames[0]]

gold_list = []
else_list = []

one = True
for row in sheet.rows:
    if one:
        one = False
        continue

    if row[1].value > 4:
        temp = []
        for cell in row:
            temp.append(cell.value)
        gold_list.append(temp)
    else:
        temp = []
        for cell in row:
            temp.append(cell.value)
        else_list.append(temp)

# print(gold_list)
# print(else_list)

sr_sheet = wb.create_sheet("Super rare")
for item in gold_list:
    sr_sheet.append(item)

rare_sheet = wb.create_sheet("Rare")
for item in else_list:
    rare_sheet.append(item)

wb.save(creat_path)
print("done.")
