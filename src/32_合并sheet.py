import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import config

data_path = config.config.absolute_create_files_PATH + "/practice1_azur_lane.xlsx"
creat_path = config.config.absolute_create_files_PATH + "/32_all_in_azur_lane.xlsx"


def sheet2list(sheet):
    rs = []

    for row in sheet.rows:
        temp = []

        for cell in row:
            temp.append(cell.value)

        rs.append(temp)

    return rs


def conbine_sheet(sheetlist, target_sheet: Worksheet):

    for sheet in sheetlist:
        value_list = sheet2list(sheet)
        for row in value_list:
            target_sheet.append(row)


wb = openpyxl.load_workbook(data_path)
sheet_list = []
for i in range(1,3):
    sheet_list.append(wb[wb.sheetnames[i]])

sheet_all = wb.create_sheet("All")

conbine_sheet(sheet_list, sheet_all)

wb.save(creat_path)
print("done")