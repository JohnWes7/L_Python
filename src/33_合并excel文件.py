import os

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import config
import openpyxl

def open_workbooks(path_list : list[str]):
    wb_list = []
    for path in path_list:
        wb = load_workbook(path)
        wb_list.append(wb)
    
    return wb_list

def all_sheets(wb : Workbook):
    all_sheet = []

    for name in wb.sheetnames:
        sheet = wb[name]
        all_sheet.append(sheet)
    
    return all_sheet

def data_of_sheet(sheet : Worksheet):
    data = []
    for row in sheet.rows:
        temp = []
        for cell in row:
            temp.append(cell.value)
        data.append(temp)
    
    return data

dir_list = os.listdir(config.config.absolute_create_files_PATH)

for i in range(dir_list.__len__()):
    dir_list[i] = config.config.absolute_create_files_PATH + "/" + dir_list[i]

wb_list = open_workbooks(dir_list)

all_wb_data = []

for wb in wb_list:
    sheet_list = all_sheets(wb)

    for sheet in sheet_list:
        value_data = data_of_sheet(sheet)
        all_wb_data.extend(value_data)
    
print(all_wb_data)

wb = Workbook()
sheet = wb[wb.sheetnames[0]]
sheet.title = "ALL info in dir"
for row in all_wb_data:
    sheet.append(row)
wb.save(config.config.absolute_create_files_PATH + "/all_in_dir.xlsx")
print("done")
    
