from xml.etree.ElementTree import indent
import config
import openpyxl

file_name = "sales_table.xlsx"

wb = openpyxl.load_workbook(config.config.absolute_data_PATH + "/" + file_name)

sheet = wb[wb.sheetnames[0]]

# 获取所有的除开开头的行
sum_list = []
for row in sheet.iter_rows(min_row=2, min_col=2):

    # 遍历一行的所有的数据 并且累加
    sum = 0
    for cell in row:
        temp = str(cell.value)
        if temp.isdigit():
            sum += cell.value

    # 累加结果存入链表
    sum_list.append(sum)

# print(sheet.max_column)

# 将数据加入
sheet.cell(1, sheet.max_column+1).value = "总计"
max_col = sheet.max_column
index = 0
for i in range(2, sheet.max_row+1):
    sheet.cell(i, max_col).value = sum_list[index]
    index += 1

wb.save(config.config.absolute_create_files_PATH+"/(out)_"+file_name)
print("done")
