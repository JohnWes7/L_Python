import openpyxl
import config
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment

# steam手机恢复吗 R46005
wb = Workbook()
sheet = wb[wb.sheetnames[0]]

font = Font(name="微软雅黑", size="30", bold=True, italic=True, color="7B68EE")
sheet["A1"] = "hello python"
sheet["A1"].font = font

# sheet.row_dimensions[1].height = 30
# sheet.column_dimensions["A"].width = 100
sheet.row_dimensions[1].height = 40
sheet.column_dimensions["A"].width = 65

sheet["C6"] = "DIO"
sheet["c6"].alignment = Alignment(horizontal="center", vertical="bottom")

sheet["h6"].fill = PatternFill(patternType="solid",fgColor="FF69B4")

wb.save(config.config.absolute_create_files_PATH+"/"+"test_font.xlsx")
print("done")
