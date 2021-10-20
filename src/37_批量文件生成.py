import os
from openpyxl.workbook.workbook import Workbook
import config
import openpyxl

wb = openpyxl.load_workbook(config.config.absolute_data_PATH + "/" + "sales_table.xlsx")

sheet = wb[wb.sheetnames[0]]

title = []
ros = []

for i,column in enumerate(sheet.columns):
    tmp = []
    if i == 0:
        for cell in column:
            title.append(cell.value)
    else:
        for cell in column:
            tmp.append(cell.value)
        ros.append(tmp)

dir_path = config.config.absolute_create_files_PATH+f'/split_tables'
if os.path.isdir(dir_path) == False:
    os.makedirs(dir_path)
    
for row in ros:
    tmp = Workbook()
    tmp_sheet = tmp[tmp.sheetnames[0]]

    tmp_sheet.append(title)
    tmp_sheet.append(row)

    tmp.save(config.config.absolute_create_files_PATH+f'/split_tables/{row[0]}.xlsx')

