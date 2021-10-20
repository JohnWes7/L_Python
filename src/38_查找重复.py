import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from config import config


def value_rows_in_sheet(sheet: Worksheet, min_row: int = 0, min_col: int = 0):
    row_list = []

    for row in sheet.iter_rows(min_row=min_row,min_col=min_col):
        temp = []
        for cell in row:
            temp.append(cell.value)
        row_list.append(temp)

    return row_list

if __name__ == "__main__":
    repeat_flag_index = 0

    wb = openpyxl.load_workbook(config.absolute_data_PATH+'/all_in_dir.xlsx')
    sheet = wb[wb.sheetnames[0]]

    value_list = value_rows_in_sheet(sheet=sheet)
    print(value_list)

    name_list = []
    reapeat_index = []
    for i,row in enumerate(value_list[1:]):
        if row[repeat_flag_index] in name_list:
            reapeat_index.append(i)
            print(i,row[repeat_flag_index])
        else:
            name_list.append(row[repeat_flag_index])

    print(name_list)
    print(reapeat_index)
    
    fill = PatternFill(patternType="solid",fgColor="FFA500")
    for index in reapeat_index:

        for row in sheet.iter_rows(min_row=index+2,max_row=index+2):
            for cell in row:
                cell.fill = fill
    
    wb.save(config.absolute_create_files_PATH+"/查找重复标记.xlsx")
    print("done")

    

