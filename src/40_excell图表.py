import openpyxl
from openpyxl import Workbook
from openpyxl.chart.line_chart import LineChart
from openpyxl.chart.pie_chart import PieChart
from openpyxl.chart.reference import Reference
from openpyxl.chart.bar_chart import BarChart
from openpyxl.worksheet.dimensions import SheetDimension
from config import config

def creat_linechart():
    filename = 'test_图表数据1.xlsx'
    wb = openpyxl.load_workbook(config.absolute_data_PATH+f'/{filename}')
    sheet = wb[wb.sheetnames[0]]

    lc = LineChart()
    lc.title = "份额表"
    lc.x_axis.title = "时间"
    lc.y_axis.title = "份额"
    lc.style = 2

    data = Reference(worksheet=sheet,min_col=1,max_col=sheet.max_column,min_row=2,max_row=sheet.max_row)
    categories = Reference(worksheet=sheet,min_col=2,max_col=sheet.max_column,min_row=1)

    #title 导入的是这个数据是每一条线的title
    lc.add_data(data=data,from_rows=True,titles_from_data=True)
    lc.set_categories(categories)

    sheet.add_chart(lc, f'A{sheet.max_row+2}')

    wb.save(config.absolute_create_files_PATH + f'/线图{filename}')


def creat_piechart():
    file_name = "test_图表数据1.xlsx"
    
    wb = openpyxl.load_workbook(config.absolute_data_PATH+f"/{file_name}")
    sheet = wb[wb.sheetnames[1]]
    
    pie_chart = PieChart()

    label = Reference(worksheet=sheet,min_col=1,min_row=2,max_row=sheet.max_row)
    data = Reference(worksheet=sheet,min_col=2,max_col=sheet.max_column,min_row=1,max_row=sheet.max_row)

    pie_chart.add_data(data=data,titles_from_data=True)
    pie_chart.set_categories(label)
    
    sheet.add_chart(pie_chart, f'A{sheet.max_row+2}')

    wb.save(config.absolute_create_files_PATH + f'/饼图{file_name}')


def creat_barchart():
    rows=[['Number','Batch 1','Batch 2'],
            ['A',10,30],
            ['B',40,60],
            ['C',50,70],
            ['D',20,10],
            ['E',10,40],
            ['F',50,30]]
    
    wb = Workbook()
    sheet = wb[wb.sheetnames[0]]

    for row in rows:
        sheet.append(row)
    

    barchart = BarChart()
    barchart.type = "col"
    barchart.title = "Bar Chart"
    barchart.y_axis.title = '数量'
    barchart.x_axis.title = '字母'

    data = Reference(worksheet=sheet,min_col=2,min_row=1,max_col=sheet.max_column,max_row=sheet.max_row)
    cats = Reference(worksheet=sheet,min_col=1,min_row=2,max_row=sheet.max_row)

    barchart.add_data(data=data,titles_from_data=True)
    barchart.set_categories(cats)

    sheet.add_chart(barchart, f"A{sheet.max_row+1}")

    wb.save(config.absolute_create_files_PATH + '/柱状图.xlsx')

if __name__ == "__main__":
    #creat_linechart()
    #creat_piechart()
    creat_barchart()

    print("done")