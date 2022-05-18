import random
import sys
from time import sleep
from numpy import Infinity
from openpyxl import Workbook
from urllib import request
import bs4
import gc

url = "https://guba.eastmoney.com/"

headtemplate = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.54 Safari/537.36 Edg/101.0.1210.39',
    'Host': "guba.eastmoney.com",
    'Referer': "https://guba.eastmoney.com/remenba.aspx?type=1&tab=1"
}

H_dataurl = "https://guba.eastmoney.com/remenba.aspx?type=1&tab=1"
S_dataurl = "https://guba.eastmoney.com/remenba.aspx?type=1&tab=2"

ip_pool_api = "http://15140409061.v4.dailiyun.com/query.txt?key=NPX029206M&word=&count=100&rand=false&ltime=0&norepeat=false&detail=false"


# 保存excel路径
save_path = "yzn.xlsx"

# proxyaddr = "代理IP地址"    #代理IP地址
# proxyport = 57114          #代理IP端口
proxyusernm = "15140409061"  # 代理帐号
proxypasswd = "15140409061"  # 代理密码


#获取所有股吧有的股票种类信息
def get_stock_list() -> list:
    st_list = []
    find_stock_data(stock_list_url=H_dataurl, stock_list=st_list)
    find_stock_data(stock_list_url=S_dataurl, stock_list=st_list)

    return st_list

# 获取ip池的ip列表 使用API


def get_all_ip():
    iplist = []
    print("开始获取ip池")
    try:
        resp = request.build_opener().open(ip_pool_api)
        text = resp.read().decode()
    except Exception as e:
        print(e)
    iplist = text.split("\r\n")
    iplist.remove("")

    return iplist

# 使用ip池 随机生成opener


def get_radom_opener():
    ip = random.sample(get_all_ip(), 1)
    ip = ip.pop(0)

    proxyurl = "http://"+proxyusernm+":"+proxypasswd+"@" + ip

    proxy = request.ProxyHandler({"http": proxyurl, "https": proxyurl})
    return request.build_opener(proxy)

# 获取单个股吧评论


def get_comment(stock: dict, num: int):
    page = 0
    com_list = []

    if num == -1:
        num = Infinity

    while com_list.__len__() < num:
        page += 1
        # 获取
        page_url = stock["href"].replace(
            ".html", "") + "_" + str(page) + ".html"
        m_request = request.Request(
            url=page_url, headers=headtemplate, method="GET")

        # 获取当前页全部html文本
        text = None

        while True:

            # 使用try catch 如果连接失败就去获取ip池的代理 直到成功
            try:
                resp = request.urlopen(m_request, timeout=5)
                text = resp.read().decode()
                break
            except Exception as e:
                print(e)
                request.install_opener(get_radom_opener())

        # 转成soup
        soup = bs4.BeautifulSoup(text, "html.parser")

        # soup查找
        result = soup.find_all(attrs={"class": "articleh"}, id=False)
        print(f"该页找到{len(result)}条数据")
        if len(result) == 0:
            print(f"已经获取了所有{len(com_list)}")
            break

        # 遍历该页全部数据
        for li in result:
            index = 1
            temp_dict = {
                "code": stock.get("code"),
                "name": stock.get("name"),
                "page": page
            }

            for item in li.contents:
                if type(item) is bs4.element.NavigableString:
                    continue

                if index == 1:
                    temp_dict["readed"] = item.string

                if index == 2:
                    temp_dict["comment"] = item.string

                if index == 3:
                    # 有个问董秘也是a
                    a = item.find(name="a", recursive=False)
                    temp_dict["title"] = a.attrs["title"]

                if index == 4:
                    if not item.a == None:
                        temp_dict["author"] = item.a.string
                    else:
                        temp_dict["author"] = item.span.string

                if index == 5:
                    temp_dict["upd"] = item.string

                index += 1

            com_list.append(temp_dict)

            if com_list.__len__() >= num:
                break

    print(len(com_list))
    return com_list

# 获取股票页面的所有股票信息


def find_stock_data(stock_list_url: str, stock_list: list):
    # 获取全部html文本
    resp = request.urlopen(url=stock_list_url)
    text = resp.read().decode()

    # 转成soup
    soup = bs4.BeautifulSoup(text, "html.parser")
    ul = soup.find(attrs={"class": "ngblistul2"})
    ul_hide = soup.find(attrs={"class": "ngblistul2 hide"})

    # 遍历存储所有股吧名称和网页以及代号
    for ul_list in [ul, ul_hide]:
        for item in ul_list.contents:
            # 如果是文本节点则跳过
            if type(item) is bs4.element.NavigableString:
                continue

            # 获取名字 连接 代号 存入outlist
            inner_text = item.a.text
            href = item.a.attrs["href"]
            temp = inner_text.split(")")
            code = temp[0]
            name = temp[1]
            code = code.replace("(", "")

            temp_dict = {
                "href": url + href,
                "code": code,
                "name": name
            }

            stock_list.append(temp_dict)

# 创造 excel Workbook


def creat_excel() -> Workbook:
    wb = Workbook()

    return wb


# sheet 添加标题头
def add_sheet_head(sheet):
    sheet.append(["readed", "comment", "title", "author",
                 "update_time", "page", "code", "name"])


# 导入excel 分开sheet


def append2excel(com_list, wb: Workbook, sheet_name: str, index, save_path):
    # 去掉非法字符
    sheet_name = sheet_name.replace("*", "")

    sheet = wb.create_sheet(sheet_name, index=index)
    add_sheet_head(sheet)

    for idct in com_list:
        sheet.append(idct)

    print("写入中请勿关闭")
    wb.save(save_path)
    print(f"写入{index} {sheet_name}完毕")


def append2sheet(com_list, sheet, save_path, wb: Workbook):
    for idct in com_list:
        sheet.append(idct)

    print("写入中请勿关闭")
    wb.save(save_path)
    print("该类股票写入完成")

# 转成excel能用的


def normalize_excel_dict(com_list):
    new_list = []
    for item in com_list:
        tempdict = {}
        tempdict["A"] = item.get("readed")
        tempdict["B"] = item.get("comment")
        tempdict["C"] = item.get("title")
        tempdict["D"] = item.get("author")
        tempdict["E"] = item.get("upd")
        tempdict["F"] = item.get("page")
        tempdict["G"] = item.get("code")
        tempdict["H"] = item.get("name")
        new_list.append(tempdict)
    return new_list


# 获得股票list里面的所有comment 并保存
def get_all_comment(select_list, is_divide_sheet, com_num, save_path):
    wb = creat_excel()
    sheet = wb[wb.sheetnames[0]]
    if not is_divide_sheet:
        add_sheet_head(sheet=sheet)

    # 遍历获取评论
    for index, stock in enumerate(select_list):
        print("开始获取"+stock.get("name"))
        com = get_comment(stock=stock, num=com_num)
        print("获取完毕")
        # 一边获取一边写入

        if is_divide_sheet:
            append2excel(com_list=normalize_excel_dict(com), wb=wb, sheet_name=stock.get(
                "name"), index=index, save_path=save_path)
        else:
            append2sheet(com_list=normalize_excel_dict(com),
                         sheet=sheet, save_path=save_path, wb=wb)
        gc.collect()


def main():
    # 设定参数
    temp = input("输入股票数量:\n")
    st_num = int(temp)
    temp = input("输入每种股票评论数量:\n")
    com_num = int(temp)
    is_divide_sheet = False
    while True:
        temp = input("是否分开sheet保存 y/n\n")
        if temp == "y":
            is_divide_sheet = True
            break
        elif temp == "n":
            is_divide_sheet = False
            break
        else:
            print("输入有误请重新输入")

    # 获取所有股票类型列表信息
    st_list = get_stock_list()

    # 随机抽取指定数量的股票
    select_list = random.sample(st_list, st_num)
    print("抽取的股票类型:")
    print(select_list)

    get_all_comment(select_list=select_list,is_divide_sheet=is_divide_sheet,com_num=com_num,save_path=save_path)

    print("done")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("错误 建议截图发给我")
        info = sys.exc_info()
        print(info)
        print(e)
        input("回车退出")
